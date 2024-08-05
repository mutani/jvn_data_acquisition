import openpyxl
import csv
from datetime import datetime, timedelta

def doExcelWriting(keywordList, dateStr):
    date = datetime.strptime(dateStr, '%Y%m%d')
    week = str((date.day - 1) // 7 + 1)
    fileName = 'output/' + week + 'W_' + '脆弱性確認' + '(' + dateStr + ')'
    date -= timedelta(days=7)
    
    # jvmDataListをJVN情報.xlsxに書き込む
    from openpyxl import load_workbook
    wb = load_workbook('脆弱性確認ひな形.xlsx')

    sheetname = ''
    rowCounter = 0
    # 検索キーワードの数だけループ
    for i in range(len(keywordList)):
        sheet = wb[keywordList[i][0][0]]
        jvmDataList = keywordList[i][2]
        
        if jvmDataList[0][0] != 0:
            jvmDataList[0].sort(key=lambda x: x[5], reverse=True)

        # JVN情報を書き込む
        for j in range(len(jvmDataList[0])):
            if jvmDataList[0][0] == 0:
                sheet.cell(row=12 + rowCounter, column=4, value=0)
                sheet.cell(row=j+14+rowCounter, column=5, value='該当するデータがありません。')
            else:
                sheet.cell(row=12 + rowCounter, column=4, value=len(jvmDataList[0]))
                for k in range(6):
                    sheet.cell(row=j+14+rowCounter, column=k+4, value=jvmDataList[0][j][k])
                    if k == 4 or k == 5:
                        jvmdt = datetime.strptime(jvmDataList[0][j][k], '%Y/%m/%d')
                        #jvmdtとdateの差分を計算し７日以内のデータを検知した場合
                        if (date - jvmdt).days <= 7:
                            #csv検知したデータを書き込む
                            with open(fileName + '.csv', 'a', newline='', encoding='utf-8') as f:
                                writer = csv.writer(f)
                                writer.writerow([jvmDataList[0][j][0], jvmDataList[0][j][1], jvmDataList[0][j][2], jvmDataList[0][j][3], jvmDataList[0][j][4], jvmDataList[0][j][5]])
            if j == 99:
                break

        if i < len(keywordList) - 1:
            if keywordList[i][0][0] == keywordList[i + 1][0][0]:
                rowCounter += len(jvmDataList[0]) + 4
            else:
                rowCounter = 0

    wb.save(fileName + '.xlsx')
